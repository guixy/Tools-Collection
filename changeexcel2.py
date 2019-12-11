#-*-coding:utf-8-*-
__author__ = 'S981325'
import re
import win32ui
import os
import xlrd
import openpyxl
def ReadAndAnalyse(filename):
    file=xlrd.open_workbook(filename)
    sheet=file.sheet_names()[0]
    sheet1=file.sheet_by_name(sheet)
    n=sheet1.nrows

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("sheet",0)
    E = sheet1.cell_value(0,3)
    E1 = re.split('\(', E)[0]


    for i in range(0,n):
        rows=sheet1.row_values(i)
        A=rows[0]
        A1=re.split(' ',A)[1]
        B=rows[1]
        B1=re.split('uint8 ',B)[1]
        B1=re.split('\[',B1)[0]
        E2 = E1 + "(" + A1 + ")]=" + B1 + "[0];"
        ws.cell(i*10+1,1,A)
        ws.cell(i*10+1, 2, B)
        ws.cell(i*10+1, 5, E2)
        for j in range(1,10):
            E2 = E1 + "(" + A1 + ")+%s]="%str(j) + B1 + "[%s];"%str(j)
            #ws.cell(i*10 + j+1, 1, A)
            #ws.cell(i*10 + j+1, 2, B)
            ws.cell(i*10 + j+1, 5, E2)

    file=re.split(os.getcwd(),filename)[0]
    wb.save(file[0:len(file)-5]+'_modify.xlsx')



dlg = win32ui.CreateFileDialog(1)  # 1表示打开文件对话框
dlg.SetOFNInitialDir(os.getcwd())  # 设置打开文件对话框中的初始显示目录
dlg.DoModal()

filename = dlg.GetPathName()  # 获取选择的文件名称

ReadAndAnalyse(filename)