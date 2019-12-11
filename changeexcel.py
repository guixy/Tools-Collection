import xlrd
import openpyxl
import win32ui
import os
import re

def ReadAndAnalyse(filename):
    file=xlrd.open_workbook(filename)
    sheet=file.sheet_names()[0]
    sheet1=file.sheet_by_name(sheet)
    n=sheet1.nrows

    wb = openpyxl.Workbook()
    ws = wb.create_sheet("sheet",0)


    for i in range(0,n):
        rows=sheet1.row_values(i)
        A=rows[0]
        A1=re.split(' ',A)[1]
        B=rows[1]
        B1=re.split('ExtendedType ',B)[1]
        E=rows[4]
        E1=re.split('\(',E)[0]
        E2=E1+"("+A1+",&"+B1[:len(B1)-1]+");"
        ws.cell(i+1,1,A)
        ws.cell(i+1, 2, B)
        ws.cell(i+1, 5, E2)

    file=re.split(os.getcwd(),filename)[0]
    wb.save(file[0:len(file)-5]+'_modify.xlsx')



dlg = win32ui.CreateFileDialog(1)  # 1表示打开文件对话框
dlg.SetOFNInitialDir(os.getcwd())  # 设置打开文件对话框中的初始显示目录
dlg.DoModal()

filename = dlg.GetPathName()  # 获取选择的文件名称

ReadAndAnalyse(filename)
#self.lineEdit_InputId_AI.setText(filename)  #将获取的文件名称写入名为“lineEdit_InputId_AI”可编辑文本框中







